from . import reports_bp
from flask import render_template, request, send_file, flash
from app.auth.utils import login_required, role_required
from app.db import get_db
import io
from datetime import date, timedelta
 
 
def get_date_range():
    date_from = request.args.get('date_from', (date.today() - timedelta(days=30)).isoformat())
    date_to   = request.args.get('date_to', date.today().isoformat())
    return date_from, date_to
 
 
def fetch_valuation():
    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute(
        '''SELECT i.item_id, i.name, c.name as category, i.unit_of_measure,
           i.current_stock, i.unit_cost,
           (i.current_stock * i.unit_cost) as total_value,
           i.status
           FROM inventory_items i
           JOIN categories c ON i.category_id = c.category_id
           WHERE i.status != %s
           ORDER BY total_value DESC''',
        ('Discontinued',)
    )
    rows = cursor.fetchall()
    cursor.close()
    return rows
 
 
def fetch_movement(date_from, date_to):
    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute(
        '''SELECT t.transaction_date, t.transaction_type, t.sub_type,
           i.name as item_name, i.unit_of_measure,
           t.quantity, t.unit_cost, t.total_value,
           t.reference_number, u.username
           FROM transactions t
           JOIN inventory_items i ON t.item_id = i.item_id
           JOIN users u ON t.performed_by = u.user_id
           WHERE t.transaction_date BETWEEN %s AND %s
           ORDER BY t.transaction_date DESC''',
        (date_from, date_to)
    )
    rows = cursor.fetchall()
    cursor.close()
    return rows
 
 
def fetch_low_stock():
    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute(
        '''SELECT i.item_id, i.name, c.name as category,
           i.current_stock, i.reorder_level, i.unit_of_measure,
           i.unit_cost, s.name as supplier_name
           FROM inventory_items i
           JOIN categories c ON i.category_id = c.category_id
           LEFT JOIN suppliers s ON i.supplier_id = s.supplier_id
           WHERE i.current_stock <= i.reorder_level AND i.status = %s
           ORDER BY i.current_stock ASC''',
        ('Active',)
    )
    rows = cursor.fetchall()
    cursor.close()
    return rows
 
 
@reports_bp.route('/')
@login_required
@role_required('Admin', 'Manager')
def index():
    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute(
        '''SELECT c.name as category,
           COALESCE(SUM(i.current_stock * i.unit_cost), 0) as value
           FROM categories c
           LEFT JOIN inventory_items i ON c.category_id = i.category_id AND i.status != %s
           GROUP BY c.category_id, c.name
           HAVING value > 0
           ORDER BY value DESC''',
        ('Discontinued',)
    )
    category_data = cursor.fetchall()
    cursor.execute(
        '''SELECT
           DATE_FORMAT(transaction_date, '%b %Y') as month,
           YEAR(transaction_date) as yr,
           MONTH(transaction_date) as mo,
           SUM(CASE WHEN transaction_type='Inflow' THEN total_value ELSE 0 END) as inflow,
           SUM(CASE WHEN transaction_type='Outflow' THEN total_value ELSE 0 END) as outflow
           FROM transactions
           WHERE transaction_date >= DATE_SUB(CURDATE(), INTERVAL 6 MONTH)
           GROUP BY yr, mo, month
           ORDER BY yr, mo'''
    )
    monthly_rows = cursor.fetchall()
    cursor.close()
    monthly_data = {
        'labels':  [r['month'] for r in monthly_rows],
        'inflow':  [float(r['inflow']) for r in monthly_rows],
        'outflow': [float(r['outflow']) for r in monthly_rows],
    }
    return render_template('reports/index.html',
        category_data=category_data,
        monthly_data=monthly_data)

 
 
@reports_bp.route('/valuation')
@login_required
@role_required('Admin', 'Manager')
def valuation():
    rows = fetch_valuation()
    total = sum(float(r['total_value']) for r in rows)
    return render_template('reports/valuation.html', rows=rows, total=total)
 
 
@reports_bp.route('/movement')
@login_required
@role_required('Admin', 'Manager')
def movement():
    date_from, date_to = get_date_range()
    rows = fetch_movement(date_from, date_to)
    inflow_total  = sum(float(r['total_value']) for r in rows if r['transaction_type'] == 'Inflow')
    outflow_total = sum(float(r['total_value']) for r in rows if r['transaction_type'] == 'Outflow')
    return render_template('reports/movement.html',
        rows=rows, date_from=date_from, date_to=date_to,
        inflow_total=inflow_total, outflow_total=outflow_total)
 
 
@reports_bp.route('/low-stock')
@login_required
@role_required('Admin', 'Manager')
def low_stock():
    rows = fetch_low_stock()
    return render_template('reports/low_stock.html', rows=rows)
 
 
@reports_bp.route('/export/excel/<report_type>')
@login_required
@role_required('Admin', 'Manager')
def export_excel(report_type):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
 
    wb = openpyxl.Workbook()
    ws = wb.active
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
 
    if report_type == 'valuation':
        ws.title = 'Inventory Valuation'
        headers = ['Item ID','Name','Category','Unit','Stock','Unit Cost (KES)','Total Value (KES)','Status']
        rows = fetch_valuation()
        ws.append(headers)
        for r in rows:
            ws.append([r['item_id'], r['name'], r['category'], r['unit_of_measure'],
                r['current_stock'], float(r['unit_cost']), float(r['total_value']), r['status']])
 
    elif report_type == 'movement':
        date_from, date_to = get_date_range()
        ws.title = 'Stock Movement'
        headers = ['Date','Item','Type','Sub-Type','Qty','Unit','Unit Cost','Total Value','Reference','Recorded By']
        rows = fetch_movement(date_from, date_to)
        ws.append(headers)
        for r in rows:
            ws.append([str(r['transaction_date']), r['item_name'], r['transaction_type'],
                r['sub_type'], r['quantity'], r['unit_of_measure'],
                float(r['unit_cost']), float(r['total_value']),
                r['reference_number'] or '', r['username']])
 
    elif report_type == 'low_stock':
        ws.title = 'Low Stock Report'
        headers = ['Item ID','Name','Category','Current Stock','Reorder Level','Unit','Unit Cost','Supplier']
        rows = fetch_low_stock()
        ws.append(headers)
        for r in rows:
            ws.append([r['item_id'], r['name'], r['category'],
                r['current_stock'], r['reorder_level'], r['unit_of_measure'],
                float(r['unit_cost']), r['supplier_name'] or 'N/A'])
    else:
        flash('Invalid report type.', 'danger')
        return
 
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
 
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
 
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'buildright_{report_type}.xlsx')
 
 
@reports_bp.route('/export/pdf/<report_type>')
@login_required
@role_required('Admin', 'Manager')
def export_pdf(report_type):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
 
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm,
        leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []
 
    title_map = {
        'valuation': 'Inventory Valuation Report',
        'movement':  'Stock Movement Report',
        'low_stock': 'Low Stock Report',
    }
    elements.append(Paragraph(title_map.get(report_type, 'Report'), styles['Title']))
    elements.append(Paragraph(f'Generated: {date.today().strftime("%d %B %Y")}', styles['Normal']))
    elements.append(Spacer(1, 0.4*cm))
 
    header_color = colors.HexColor('#2C3E50')
    row_alt      = colors.HexColor('#ECF0F1')
 
    if report_type == 'valuation':
        rows = fetch_valuation()
        data = [['Item ID','Name','Category','Unit','Stock','Unit Cost','Total Value','Status']]
        for r in rows:
            data.append([r['item_id'], r['name'][:30], r['category'], r['unit_of_measure'],
                str(r['current_stock']), f'KES {float(r["unit_cost"]):,.2f}',
                f'KES {float(r["total_value"]):,.2f}', r['status']])
 
    elif report_type == 'movement':
        date_from, date_to = get_date_range()
        rows = fetch_movement(date_from, date_to)
        data = [['Date','Item','Type','Sub-Type','Qty','Unit Cost','Total Value','By']]
        for r in rows:
            data.append([str(r['transaction_date']), r['item_name'][:25],
                r['transaction_type'], r['sub_type'], str(r['quantity']),
                f'KES {float(r["unit_cost"]):,.2f}', f'KES {float(r["total_value"]):,.2f}',
                r['username']])
 
    elif report_type == 'low_stock':
        rows = fetch_low_stock()
        data = [['Item ID','Name','Category','Current Stock','Reorder Level','Unit','Supplier']]
        for r in rows:
            data.append([r['item_id'], r['name'][:30], r['category'],
                str(r['current_stock']), str(r['reorder_level']),
                r['unit_of_measure'], r['supplier_name'] or 'N/A'])
    else:
        flash('Invalid report type.', 'danger')
        return
 
    t = Table(data, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), header_color),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, row_alt]),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#BDC3C7')),
        ('ALIGN',      (0,0), (-1,-1), 'LEFT'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING',    (0,0), (-1,-1), 5),
    ])
    t.setStyle(style)
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
        as_attachment=True, download_name=f'buildright_{report_type}.pdf')
